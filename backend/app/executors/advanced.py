"""高级模块执行器实现 - 异步版本"""
import asyncio
import base64
import json
import random
import time
from pathlib import Path

from .base import (
    ModuleExecutor,
    ExecutionContext,
    ModuleResult,
    register_executor,
)
from .type_utils import to_int, to_float


@register_executor
class ApiRequestExecutor(ModuleExecutor):
    """API请求模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "api_request"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import httpx
        
        request_url = context.resolve_value(config.get('requestUrl', ''))
        request_method = config.get('requestMethod', 'GET').upper()
        request_headers_str = context.resolve_value(config.get('requestHeaders', ''))
        request_body_str = context.resolve_value(config.get('requestBody', ''))
        variable_name = config.get('variableName', '')
        request_timeout = to_int(config.get('requestTimeout', 30), 30, context)
        
        if not request_url:
            return ModuleResult(success=False, error="请求地址不能为空")
        
        try:
            headers = {}
            if request_headers_str:
                try:
                    headers = json.loads(request_headers_str)
                except json.JSONDecodeError as e:
                    return ModuleResult(success=False, error=f"请求头JSON格式错误: {str(e)}")
            
            body = None
            if request_body_str:
                try:
                    body = json.loads(request_body_str)
                except json.JSONDecodeError:
                    body = request_body_str
            
            async with httpx.AsyncClient(timeout=request_timeout) as client:
                response = await client.request(
                    method=request_method,
                    url=request_url,
                    headers=headers,
                    json=body if isinstance(body, dict) else None,
                    data=body if isinstance(body, str) else None,
                )
            
            try:
                response_data = response.json()
            except:
                response_data = response.text
            
            if variable_name:
                context.set_variable(variable_name, response_data)
            
            display_content = str(response_data)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(
                success=True,
                message=f"请求成功 ({response.status_code}): {display_content}",
                data={'status_code': response.status_code, 'response': response_data}
            )
        
        except httpx.TimeoutException:
            return ModuleResult(success=False, error=f"请求超时 ({request_timeout}秒)")
        except httpx.ConnectError:
            return ModuleResult(success=False, error="无法连接到服务器，请检查网络和URL")
        except Exception as e:
            return ModuleResult(success=False, error=f"API请求失败: {str(e)}")


@register_executor
class JsonParseExecutor(ModuleExecutor):
    """JSON解析模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "json_parse"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_variable = config.get('sourceVariable', '')
        json_path = config.get('jsonPath', '')
        variable_name = config.get('variableName', '')
        column_name = config.get('columnName', '')
        
        if not source_variable:
            return ModuleResult(success=False, error="源数据变量不能为空")
        
        if not json_path:
            return ModuleResult(success=False, error="JSONPath表达式不能为空")
        
        source_data = context.get_variable(source_variable)
        if source_data is None:
            return ModuleResult(success=False, error=f"变量 '{source_variable}' 不存在")
        
        if isinstance(source_data, str):
            try:
                source_data = json.loads(source_data)
            except json.JSONDecodeError as e:
                return ModuleResult(success=False, error=f"源数据不是有效的JSON: {str(e)}")
        
        try:
            result = self._parse_jsonpath(source_data, json_path)
            
            if variable_name:
                context.set_variable(variable_name, result)
            
            if column_name:
                context.add_data_value(column_name, result)
            
            display_content = str(result)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(success=True, message=f"解析成功: {display_content}", data=result)
        
        except Exception as e:
            return ModuleResult(success=False, error=f"JSON解析失败: {str(e)}")
    
    def _parse_jsonpath(self, data, path: str):
        """简单的JSONPath解析器"""
        if path.startswith('$'):
            path = path[1:]
        if path.startswith('.'):
            path = path[1:]
        
        if not path:
            return data
        
        current = data
        parts = self._split_path(path)
        
        for part in parts:
            if current is None:
                return None
            
            if part.startswith('[') and part.endswith(']'):
                index_str = part[1:-1]
                if index_str == '*':
                    if isinstance(current, list):
                        return current
                    else:
                        return None
                else:
                    try:
                        index = int(index_str)
                        if isinstance(current, list) and -len(current) <= index < len(current):
                            current = current[index]
                        else:
                            return None
                    except ValueError:
                        return None
            else:
                if '[' in part:
                    prop_name = part[:part.index('[')]
                    array_part = part[part.index('['):]
                    
                    if isinstance(current, dict) and prop_name in current:
                        current = current[prop_name]
                        current = self._parse_jsonpath(current, array_part)
                    else:
                        return None
                else:
                    if isinstance(current, dict) and part in current:
                        current = current[part]
                    else:
                        return None
        
        return current
    
    def _split_path(self, path: str) -> list:
        """分割JSONPath路径"""
        parts = []
        current = ''
        in_bracket = False
        
        for char in path:
            if char == '[':
                if current:
                    parts.append(current)
                    current = ''
                in_bracket = True
                current = '['
            elif char == ']':
                current += ']'
                parts.append(current)
                current = ''
                in_bracket = False
            elif char == '.' and not in_bracket:
                if current:
                    parts.append(current)
                    current = ''
            else:
                current += char
        
        if current:
            parts.append(current)
        
        return parts


@register_executor
class Base64Executor(ModuleExecutor):
    """Base64处理模块执行器"""

    @property
    def module_type(self) -> str:
        return "base64"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        operation = config.get("operation", "encode")
        variable_name = config.get("variableName", "")

        try:
            if operation == "encode":
                input_text = context.resolve_value(config.get("inputText", ""))
                if not input_text:
                    return ModuleResult(success=False, error="输入文本不能为空")

                result = base64.b64encode(input_text.encode("utf-8")).decode("utf-8")
                if variable_name:
                    context.set_variable(variable_name, result)

                display = result[:50] + "..." if len(result) > 50 else result
                return ModuleResult(success=True, message=f"编码成功: {display}", data=result)

            elif operation == "decode":
                input_base64 = context.resolve_value(config.get("inputBase64", ""))
                if not input_base64:
                    return ModuleResult(success=False, error="Base64字符串不能为空")

                if "," in input_base64:
                    input_base64 = input_base64.split(",", 1)[1]

                result = base64.b64decode(input_base64).decode("utf-8")
                if variable_name:
                    context.set_variable(variable_name, result)

                display = result[:50] + "..." if len(result) > 50 else result
                return ModuleResult(success=True, message=f"解码成功: {display}", data=result)

            elif operation == "file_to_base64":
                file_path = context.resolve_value(config.get("filePath", ""))
                if not file_path:
                    return ModuleResult(success=False, error="文件路径不能为空")

                path = Path(file_path)
                if not path.exists():
                    return ModuleResult(success=False, error=f"文件不存在: {file_path}")

                with open(path, "rb") as f:
                    file_data = f.read()

                result = base64.b64encode(file_data).decode("utf-8")

                ext = path.suffix.lower()
                mime_types = {
                    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                    ".gif": "image/gif", ".webp": "image/webp", ".svg": "image/svg+xml",
                    ".pdf": "application/pdf", ".txt": "text/plain",
                    ".json": "application/json", ".xml": "application/xml",
                }
                mime_type = mime_types.get(ext, "application/octet-stream")
                data_url = f"data:{mime_type};base64,{result}"

                if variable_name:
                    context.set_variable(variable_name, data_url)

                return ModuleResult(
                    success=True,
                    message=f"文件转换成功: {path.name} ({len(file_data)} 字节)",
                    data=data_url,
                )

            elif operation == "base64_to_file":
                input_base64 = context.resolve_value(config.get("inputBase64", ""))
                output_path = context.resolve_value(config.get("outputPath", ""))
                file_name = context.resolve_value(config.get("fileName", "output.bin"))

                if not input_base64:
                    return ModuleResult(success=False, error="Base64字符串不能为空")
                if not output_path:
                    return ModuleResult(success=False, error="保存路径不能为空")

                if "," in input_base64:
                    input_base64 = input_base64.split(",", 1)[1]

                file_data = base64.b64decode(input_base64)

                output_dir = Path(output_path)
                output_dir.mkdir(parents=True, exist_ok=True)

                full_path = output_dir / file_name
                with open(full_path, "wb") as f:
                    f.write(file_data)

                result_path = str(full_path)
                if variable_name:
                    context.set_variable(variable_name, result_path)

                return ModuleResult(
                    success=True,
                    message=f"文件保存成功: {full_path} ({len(file_data)} 字节)",
                    data=result_path,
                )

            else:
                return ModuleResult(success=False, error=f"未知操作类型: {operation}")

        except Exception as e:
            return ModuleResult(success=False, error=f"Base64处理失败: {str(e)}")


@register_executor
class SelectDropdownExecutor(ModuleExecutor):
    """下拉框选择模块执行器"""

    @property
    def module_type(self) -> str:
        return "select_dropdown"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        select_by = config.get('selectBy', 'value')
        value = context.resolve_value(config.get('value', ''))
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            # 先等待页面加载完成
            try:
                await context.page.wait_for_load_state('domcontentloaded', timeout=10000)
            except:
                pass
            
            await context.page.wait_for_selector(selector, state='visible', timeout=30000)
            element = context.page.locator(selector)
            
            if select_by == 'value':
                await element.select_option(value=value)
            elif select_by == 'label':
                await element.select_option(label=value)
            elif select_by == 'index':
                await element.select_option(index=int(value))
            
            return ModuleResult(success=True, message=f"已选择: {value}")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"选择下拉框失败: {str(e)}")


@register_executor
class SetCheckboxExecutor(ModuleExecutor):
    """设置复选框模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "set_checkbox"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        checked = config.get('checked', True)
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            await context.page.wait_for_selector(selector, state='visible')
            element = context.page.locator(selector)
            
            if checked:
                await element.check()
            else:
                await element.uncheck()
            
            return ModuleResult(success=True, message=f"复选框已{'勾选' if checked else '取消勾选'}")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"设置复选框失败: {str(e)}")


@register_executor
class DragElementExecutor(ModuleExecutor):
    """拖拽元素模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "drag_element"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        source_selector = context.resolve_value(config.get('sourceSelector', ''))
        target_selector = context.resolve_value(config.get('targetSelector', ''))
        target_position = config.get('targetPosition')
        
        if not source_selector:
            return ModuleResult(success=False, error="源元素选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            source = context.page.locator(source_selector)
            
            if target_selector:
                target = context.page.locator(target_selector)
                await source.drag_to(target)
            elif target_position:
                box = await source.bounding_box()
                if box:
                    start_x = box['x'] + box['width'] / 2
                    start_y = box['y'] + box['height'] / 2
                    end_x = target_position.get('x', start_x)
                    end_y = target_position.get('y', start_y)
                    
                    await context.page.mouse.move(start_x, start_y)
                    await context.page.mouse.down()
                    await context.page.mouse.move(end_x, end_y, steps=10)
                    await context.page.mouse.up()
            
            return ModuleResult(success=True, message="拖拽完成")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"拖拽元素失败: {str(e)}")


@register_executor
class ScrollPageExecutor(ModuleExecutor):
    """滚动页面模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "scroll_page"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        direction = config.get('direction', 'down')
        distance = to_int(config.get('distance', 500), 500, context)
        selector = context.resolve_value(config.get('selector', ''))
        scroll_mode = config.get('scrollMode', 'auto')  # auto, wheel, script
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            delta_x = 0
            delta_y = 0
            
            if direction == 'down':
                delta_y = distance
            elif direction == 'up':
                delta_y = -distance
            elif direction == 'right':
                delta_x = distance
            elif direction == 'left':
                delta_x = -distance
            
            # 使用鼠标滚轮模拟滚动（对抖音等虚拟滚动页面更有效）
            if scroll_mode == 'wheel' or scroll_mode == 'auto':
                try:
                    if selector:
                        # 先定位到元素中心
                        element = context.page.locator(selector).first
                        box = await element.bounding_box()
                        if box:
                            center_x = box['x'] + box['width'] / 2
                            center_y = box['y'] + box['height'] / 2
                            await context.page.mouse.move(center_x, center_y)
                            await context.page.mouse.wheel(delta_x, delta_y)
                        else:
                            raise Exception("无法获取元素位置")
                    else:
                        # 在页面中心滚动
                        viewport = context.page.viewport_size
                        if viewport:
                            await context.page.mouse.move(viewport['width'] / 2, viewport['height'] / 2)
                        await context.page.mouse.wheel(delta_x, delta_y)
                    
                    return ModuleResult(success=True, message=f"已滚动 {direction} {distance}px (鼠标滚轮)")
                except Exception as wheel_error:
                    if scroll_mode == 'wheel':
                        raise wheel_error
                    # auto 模式下，滚轮失败则尝试脚本滚动
            
            # 使用 JavaScript 滚动
            if selector:
                await context.page.locator(selector).evaluate(
                    f"el => el.scrollBy({delta_x}, {delta_y})"
                )
            else:
                await context.page.evaluate(f"window.scrollBy({delta_x}, {delta_y})")
            
            return ModuleResult(success=True, message=f"已滚动 {direction} {distance}px")
        
        except Exception as e:
            return ModuleResult(success=False, error=f"滚动页面失败: {str(e)}")


@register_executor
class UploadFileExecutor(ModuleExecutor):
    """上传文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "upload_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get("selector", ""))
        file_path = context.resolve_value(config.get("filePath", ""))

        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")

        if not file_path:
            return ModuleResult(success=False, error="文件路径不能为空")

        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")

        try:
            await context.switch_to_latest_page()

            element = context.page.locator(selector)
            tag_name = await element.evaluate("el => el.tagName.toLowerCase()")
            input_type = await element.evaluate("el => el.type || ''")

            if tag_name == "input" and input_type == "file":
                await context.page.set_input_files(selector, file_path)
            else:
                try:
                    async with context.page.expect_file_chooser(timeout=5000) as fc_info:
                        await element.click()
                    file_chooser = await fc_info.value
                    await file_chooser.set_files(file_path)
                except Exception:
                    file_inputs = context.page.locator('input[type="file"]:not([webkitdirectory])')
                    count = await file_inputs.count()
                    if count > 0:
                        await file_inputs.last.set_input_files(file_path)
                    else:
                        all_inputs = context.page.locator('input[type="file"]')
                        if await all_inputs.count() > 0:
                            await all_inputs.first.set_input_files(file_path)
                        else:
                            raise Exception("未找到文件上传入口")

            return ModuleResult(success=True, message=f"已上传文件: {file_path}")

        except Exception as e:
            return ModuleResult(success=False, error=f"上传文件失败: {str(e)}")


@register_executor
class DownloadFileExecutor(ModuleExecutor):
    """下载文件模块执行器"""

    @property
    def module_type(self) -> str:
        return "download_file"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import httpx
        import os
        from urllib.parse import urlparse, unquote
        
        download_mode = config.get("downloadMode", "click")
        trigger_selector = context.resolve_value(config.get("triggerSelector", ""))
        download_url = context.resolve_value(config.get("downloadUrl", ""))
        save_path = context.resolve_value(config.get("savePath", ""))
        file_name = context.resolve_value(config.get("fileName", ""))
        variable_name = config.get("variableName", "")

        try:
            if download_mode == "url":
                if not download_url:
                    return ModuleResult(success=False, error="下载URL不能为空")

                if not file_name:
                    parsed = urlparse(download_url)
                    file_name = unquote(os.path.basename(parsed.path)) or "downloaded_file"

                if save_path:
                    Path(save_path).mkdir(parents=True, exist_ok=True)
                    final_path = Path(save_path) / file_name
                else:
                    downloads_dir = Path("downloads")
                    downloads_dir.mkdir(exist_ok=True)
                    final_path = downloads_dir / file_name

                async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
                    response = await client.get(download_url)
                    response.raise_for_status()

                    with open(final_path, "wb") as f:
                        f.write(response.content)

                if variable_name:
                    context.set_variable(variable_name, str(final_path))

                return ModuleResult(
                    success=True,
                    message=f"已下载文件: {final_path}",
                    data=str(final_path),
                )

            else:
                if not trigger_selector:
                    return ModuleResult(success=False, error="触发元素选择器不能为空")

                if context.page is None:
                    return ModuleResult(success=False, error="没有打开的页面")

                await context.switch_to_latest_page()

                async with context.page.expect_download() as download_info:
                    await context.page.click(trigger_selector)

                download = await download_info.value

                if save_path:
                    Path(save_path).mkdir(parents=True, exist_ok=True)
                    if file_name:
                        final_path = Path(save_path) / file_name
                    else:
                        final_path = Path(save_path) / download.suggested_filename
                    await download.save_as(str(final_path))
                else:
                    final_path = await download.path()

                if variable_name:
                    context.set_variable(variable_name, str(final_path))

                return ModuleResult(
                    success=True,
                    message=f"已下载文件: {final_path}",
                    data=str(final_path),
                )

        except Exception as e:
            return ModuleResult(success=False, error=f"下载文件失败: {str(e)}")


@register_executor
class SaveImageExecutor(ModuleExecutor):
    """保存图片模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "save_image"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        selector = context.resolve_value(config.get('selector', ''))
        save_path = context.resolve_value(config.get('savePath', ''))
        variable_name = config.get('variableName', '')
        
        if not selector:
            return ModuleResult(success=False, error="选择器不能为空")
        
        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")
        
        try:
            await context.switch_to_latest_page()
            
            element = context.page.locator(selector)
            src = await element.get_attribute('src')
            
            if src and src.startswith('data:'):
                header, data = src.split(',', 1)
                image_data = base64.b64decode(data)
            else:
                image_data = await element.screenshot()
            
            if save_path:
                Path(save_path).parent.mkdir(parents=True, exist_ok=True)
                with open(save_path, 'wb') as f:
                    f.write(image_data)
                final_path = save_path
            else:
                import tempfile
                with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
                    f.write(image_data)
                    final_path = f.name
            
            if variable_name:
                context.set_variable(variable_name, final_path)
            
            return ModuleResult(success=True, message=f"已保存图片: {final_path}", data=final_path)
        
        except Exception as e:
            return ModuleResult(success=False, error=f"保存图片失败: {str(e)}")


@register_executor
class SendEmailExecutor(ModuleExecutor):
    """发送邮件模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "send_email"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        sender_email = context.resolve_value(config.get('senderEmail', ''))
        auth_code = context.resolve_value(config.get('authCode', ''))
        recipient_email = context.resolve_value(config.get('recipientEmail', ''))
        email_subject = context.resolve_value(config.get('emailSubject', ''))
        email_content = context.resolve_value(config.get('emailContent', ''))
        
        if not sender_email:
            return ModuleResult(success=False, error="发件人邮箱不能为空")
        
        if not auth_code:
            return ModuleResult(success=False, error="授权码不能为空")
        
        if not recipient_email:
            return ModuleResult(success=False, error="收件人邮箱不能为空")
        
        try:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = recipient_email
            msg['Subject'] = email_subject or '(无标题)'
            
            msg.attach(MIMEText(email_content or '', 'plain', 'utf-8'))
            
            # 使用线程池执行同步SMTP操作
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, self._send_email_sync, sender_email, auth_code, recipient_email, msg)
            
            return ModuleResult(
                success=True, 
                message=f"邮件已发送至: {recipient_email}",
                data={'recipient': recipient_email, 'subject': email_subject}
            )
        
        except smtplib.SMTPAuthenticationError:
            return ModuleResult(success=False, error="邮箱认证失败，请检查邮箱地址和授权码")
        except Exception as e:
            return ModuleResult(success=False, error=f"发送邮件失败: {str(e)}")
    
    def _send_email_sync(self, sender_email, auth_code, recipient_email, msg):
        import smtplib
        server = smtplib.SMTP_SSL('smtp.qq.com', 465)
        server.login(sender_email, auth_code)
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()


@register_executor
class ReadExcelExecutor(ModuleExecutor):
    """Excel文件读取模块执行器"""
    
    @property
    def module_type(self) -> str:
        return "read_excel"
    
    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import openpyxl
        from app.api.data_assets import get_asset_by_name
        
        file_name = context.resolve_value(config.get('fileName', ''))
        sheet_name = context.resolve_value(config.get('sheetName', ''))
        read_mode = config.get('readMode', 'cell')
        cell_address = context.resolve_value(config.get('cellAddress', ''))
        row_index = to_int(config.get('rowIndex', 1), 1, context)
        column_index = context.resolve_value(config.get('columnIndex', ''))
        start_cell = context.resolve_value(config.get('startCell', ''))
        end_cell = context.resolve_value(config.get('endCell', ''))
        variable_name = config.get('variableName', '')
        start_row = to_int(config.get('startRow', 2), 2, context)
        start_col = context.resolve_value(config.get('startCol', ''))
        
        if not file_name:
            return ModuleResult(success=False, error="请选择要读取的Excel文件")
        
        if not variable_name:
            return ModuleResult(success=False, error="请指定存储变量名")
        
        asset = get_asset_by_name(file_name)
        if not asset:
            return ModuleResult(success=False, error=f"文件 '{file_name}' 不存在")
        
        file_path = asset['path']
        is_xls = file_path.lower().endswith('.xls')
        
        try:
            # 使用线程池执行同步Excel操作
            loop = asyncio.get_event_loop()
            if is_xls:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xls, file_path, sheet_name, read_mode, 
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            else:
                result, result_type = await loop.run_in_executor(
                    None, self._read_xlsx, file_path, sheet_name, read_mode,
                    cell_address, row_index, column_index, start_cell, end_cell, start_row, start_col
                )
            
            context.set_variable(variable_name, result)
            
            display_content = str(result)
            if len(display_content) > 100:
                display_content = display_content[:100] + '...'
            
            return ModuleResult(
                success=True,
                message=f"已读取Excel数据: {display_content}",
                data={'value': result, 'type': result_type, 'file': file_name}
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"读取Excel失败: {str(e)}")
    
    def _read_xlsx(self, file_path, sheet_name, read_mode, cell_address, row_index, 
                   column_index, start_cell, end_cell, start_row, start_col):
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                wb.close()
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                wb.close()
                raise Exception("单元格模式需要指定单元格地址")
            cell = ws[cell_address]
            result = cell.value
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                wb.close()
                raise Exception("行模式需要指定有效的行号")
            row_data = []
            start_col_idx = 1
            if start_col:
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = openpyxl.utils.column_index_from_string(start_col)
                else:
                    start_col_idx = int(start_col)
            for cell in ws[row_index]:
                if cell.column >= start_col_idx:
                    row_data.append(cell.value)
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                wb.close()
                raise Exception("列模式需要指定列号或列字母")
            col_data = []
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = openpyxl.utils.column_index_from_string(col_idx)
            else:
                col_idx = int(col_idx)
            for row in ws.iter_rows(min_row=start_row, min_col=col_idx, max_col=col_idx):
                col_data.append(row[0].value)
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                wb.close()
                raise Exception("范围模式需要指定起始和结束单元格")
            range_data = []
            for row in ws[f"{start_cell}:{end_cell}"]:
                row_data = [cell.value for cell in row]
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        wb.close()
        return result, result_type
    
    def _read_xls(self, file_path, sheet_name, read_mode, cell_address, row_index,
                  column_index, start_cell, end_cell, start_row, start_col):
        import xlrd
        
        wb = xlrd.open_workbook(file_path)
        
        if sheet_name:
            if sheet_name not in wb.sheet_names():
                raise Exception(f"工作表 '{sheet_name}' 不存在")
            ws = wb.sheet_by_name(sheet_name)
        else:
            ws = wb.sheet_by_index(0)
        
        result = None
        result_type = 'unknown'
        
        if read_mode == 'cell':
            if not cell_address:
                raise Exception("单元格模式需要指定单元格地址")
            col_idx, row_idx = self._parse_cell_address(cell_address)
            result = ws.cell_value(row_idx, col_idx)
            result_type = 'cell'
        
        elif read_mode == 'row':
            if row_index is None or row_index < 1:
                raise Exception("行模式需要指定有效的行号")
            row_data = ws.row_values(row_index - 1)
            if start_col:
                start_col_idx = 0
                if isinstance(start_col, str) and start_col.isalpha():
                    start_col_idx = self._col_letter_to_index(start_col)
                else:
                    start_col_idx = int(start_col) - 1
                row_data = row_data[start_col_idx:]
            result = row_data
            result_type = 'array'
        
        elif read_mode == 'column':
            if not column_index:
                raise Exception("列模式需要指定列号或列字母")
            col_idx = column_index
            if isinstance(col_idx, str) and col_idx.isalpha():
                col_idx = self._col_letter_to_index(col_idx)
            else:
                col_idx = int(col_idx) - 1
            col_data = ws.col_values(col_idx, start_rowx=start_row - 1)
            result = col_data
            result_type = 'array'
        
        elif read_mode == 'range':
            if not start_cell or not end_cell:
                raise Exception("范围模式需要指定起始和结束单元格")
            start_col_idx, start_row_idx = self._parse_cell_address(start_cell)
            end_col_idx, end_row_idx = self._parse_cell_address(end_cell)
            range_data = []
            for r in range(start_row_idx, end_row_idx + 1):
                row_data = []
                for c in range(start_col_idx, end_col_idx + 1):
                    row_data.append(ws.cell_value(r, c))
                range_data.append(row_data)
            result = range_data
            result_type = 'matrix'
        
        return result, result_type
    
    def _col_letter_to_index(self, col_str: str) -> int:
        result = 0
        for c in col_str.upper():
            result = result * 26 + (ord(c) - ord('A') + 1)
        return result - 1
    
    def _parse_cell_address(self, address: str) -> tuple:
        col_str, row_str = '', ''
        for c in address:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        return self._col_letter_to_index(col_str), int(row_str) - 1


@register_executor
class SetClipboardExecutor(ModuleExecutor):
    """设置剪贴板模块执行器"""

    @property
    def module_type(self) -> str:
        return "set_clipboard"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import subprocess
        
        content_type = config.get("contentType", "text")
        text_content = context.resolve_value(config.get("textContent", ""))
        image_path = context.resolve_value(config.get("imagePath", ""))

        try:
            if content_type == "image":
                if not image_path:
                    return ModuleResult(success=False, error="图片路径不能为空")

                if not Path(image_path).exists():
                    return ModuleResult(success=False, error=f"图片文件不存在: {image_path}")

                try:
                    from PIL import Image
                    import tempfile

                    img = Image.open(image_path)
                    if img.mode == "RGBA":
                        img = img.convert("RGB")
                    
                    with tempfile.NamedTemporaryFile(suffix='.bmp', delete=False) as tmp:
                        tmp_path = tmp.name
                        img.save(tmp_path, "BMP")
                    
                    ps_script = f'''
Add-Type -AssemblyName System.Windows.Forms
$image = [System.Drawing.Image]::FromFile("{tmp_path.replace(chr(92), '/')}")
[System.Windows.Forms.Clipboard]::SetImage($image)
$image.Dispose()
'''
                    loop = asyncio.get_event_loop()
                    result = await loop.run_in_executor(
                        None, 
                        lambda: subprocess.run(["powershell", "-Command", ps_script], capture_output=True, text=True)
                    )
                    
                    try:
                        Path(tmp_path).unlink()
                    except:
                        pass
                    
                    if result.returncode != 0:
                        return ModuleResult(success=False, error=f"设置剪贴板失败: {result.stderr}")

                    return ModuleResult(success=True, message=f"已将图片复制到剪贴板: {image_path}")

                except ImportError:
                    return ModuleResult(success=False, error="需要安装 Pillow 库")

            else:
                if not text_content:
                    return ModuleResult(success=False, error="文本内容不能为空")

                escaped_text = text_content.replace("'", "''")
                ps_script = f"Set-Clipboard -Value '{escaped_text}'"
                
                loop = asyncio.get_event_loop()
                result = await loop.run_in_executor(
                    None,
                    lambda: subprocess.run(["powershell", "-Command", ps_script], capture_output=True, text=True)
                )
                
                if result.returncode != 0:
                    return ModuleResult(success=False, error=f"设置剪贴板失败: {result.stderr}")

                display_text = text_content[:50] + "..." if len(text_content) > 50 else text_content
                return ModuleResult(success=True, message=f"已将文本复制到剪贴板: {display_text}")

        except Exception as e:
            return ModuleResult(success=False, error=f"设置剪贴板失败: {str(e)}")


@register_executor
class GetClipboardExecutor(ModuleExecutor):
    """获取剪贴板模块执行器"""

    @property
    def module_type(self) -> str:
        return "get_clipboard"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import subprocess
        
        variable_name = config.get("variableName", "")

        if not variable_name:
            return ModuleResult(success=False, error="存储变量名不能为空")

        try:
            ps_script = "Get-Clipboard -Raw"
            
            loop = asyncio.get_event_loop()
            result = await loop.run_in_executor(
                None,
                lambda: subprocess.run(["powershell", "-Command", ps_script], capture_output=True, timeout=10)
            )

            if result.returncode != 0:
                error_msg = result.stderr.decode('utf-8', errors='ignore') if result.stderr else "未知错误"
                return ModuleResult(success=False, error=f"获取剪贴板失败: {error_msg}")

            try:
                clipboard_content = result.stdout.decode('utf-8')
            except UnicodeDecodeError:
                clipboard_content = result.stdout.decode('gbk', errors='ignore')
            
            clipboard_content = clipboard_content.rstrip('\r\n') if clipboard_content else ""
            context.set_variable(variable_name, clipboard_content)

            if not clipboard_content:
                return ModuleResult(
                    success=True,
                    message=f"剪贴板为空，已设置变量 {variable_name} 为空字符串",
                    data=clipboard_content
                )

            display_text = clipboard_content[:50] + "..." if len(clipboard_content) > 50 else clipboard_content
            return ModuleResult(
                success=True,
                message=f"已获取剪贴板内容: {display_text}",
                data=clipboard_content
            )

        except subprocess.TimeoutExpired:
            return ModuleResult(success=False, error="获取剪贴板超时")
        except Exception as e:
            return ModuleResult(success=False, error=f"获取剪贴板失败: {str(e)}")


@register_executor
class KeyboardActionExecutor(ModuleExecutor):
    """键盘操作模块执行器"""

    @property
    def module_type(self) -> str:
        return "keyboard_action"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        target_type = config.get("targetType", "page")
        selector = context.resolve_value(config.get("selector", ""))
        key_sequence = context.resolve_value(config.get("keySequence", ""))
        delay = to_int(config.get("delay", 0), 0, context)

        if not key_sequence:
            return ModuleResult(success=False, error="按键序列不能为空")

        if context.page is None:
            return ModuleResult(success=False, error="没有打开的页面")

        try:
            await context.switch_to_latest_page()

            if target_type == "element" and selector:
                element = context.page.locator(selector)
                await element.focus()
                await asyncio.sleep(0.1)

            keys = key_sequence.split("+")

            if len(keys) == 1:
                await context.page.keyboard.press(keys[0])
            else:
                modifiers = keys[:-1]
                main_key = keys[-1]

                for mod in modifiers:
                    await context.page.keyboard.down(mod)

                if delay > 0:
                    await asyncio.sleep(delay / 1000)

                await context.page.keyboard.press(main_key)

                for mod in reversed(modifiers):
                    await context.page.keyboard.up(mod)

            return ModuleResult(success=True, message=f"已执行键盘操作: {key_sequence}")

        except Exception as e:
            return ModuleResult(success=False, error=f"键盘操作失败: {str(e)}")


@register_executor
class RealMouseScrollExecutor(ModuleExecutor):
    """真实鼠标滚动模块执行器 - 使用系统级鼠标滚轮模拟"""

    @property
    def module_type(self) -> str:
        return "real_mouse_scroll"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        direction = config.get("direction", "down")
        scroll_amount = to_int(config.get("scrollAmount", 3), 3, context)  # 滚动格数
        scroll_count = to_int(config.get("scrollCount", 1), 1, context)  # 滚动次数
        scroll_interval = to_int(config.get("scrollInterval", 100), 100, context)  # 滚动间隔(毫秒)

        try:
            import ctypes
            
            # Windows API 常量
            MOUSEEVENTF_WHEEL = 0x0800
            WHEEL_DELTA = 120  # 一格滚轮的标准值
            
            # 计算滚动量（向上为正，向下为负）
            delta = WHEEL_DELTA * scroll_amount
            if direction == "down":
                delta = -delta
            
            # 执行滚动
            for i in range(scroll_count):
                # 使用 mouse_event 发送滚轮事件
                ctypes.windll.user32.mouse_event(MOUSEEVENTF_WHEEL, 0, 0, delta, 0)
                
                if i < scroll_count - 1 and scroll_interval > 0:
                    await asyncio.sleep(scroll_interval / 1000)
            
            direction_text = "向下" if direction == "down" else "向上"
            return ModuleResult(
                success=True, 
                message=f"已{direction_text}滚动 {scroll_count} 次，每次 {scroll_amount} 格"
            )

        except ImportError:
            return ModuleResult(success=False, error="此功能仅支持 Windows 系统")
        except Exception as e:
            return ModuleResult(success=False, error=f"真实鼠标滚动失败: {str(e)}")
