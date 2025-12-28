"""Excel文件资源API - 处理Excel文件上传和读取"""
import os
import uuid
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel
import openpyxl
import xlrd

router = APIRouter(prefix="/api/data-assets", tags=["data-assets"])

# 存储上传的文件信息
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)

# 内存中存储文件元数据
data_assets: dict[str, dict] = {}


class ReadExcelRequest(BaseModel):
    fileId: str
    sheetName: Optional[str] = None
    readMode: str  # 'cell', 'row', 'column', 'range'
    cellAddress: Optional[str] = None  # 如 'A1'
    rowIndex: Optional[int] = None  # 行号，从1开始
    columnIndex: Optional[int] = None  # 列号，从1开始，或列字母如'A'
    startCell: Optional[str] = None  # 范围起始，如 'A1'
    endCell: Optional[str] = None  # 范围结束，如 'C10'


@router.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    """上传Excel文件"""
    # 检查文件类型
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持Excel文件 (.xlsx, .xls)")
    
    # 生成唯一ID和文件名
    file_id = str(uuid.uuid4())
    ext = os.path.splitext(file.filename)[1]
    saved_name = f"{file_id}{ext}"
    file_path = os.path.join(UPLOAD_DIR, saved_name)
    
    # 保存文件
    content = await file.read()
    with open(file_path, 'wb') as f:
        f.write(content)
    
    # 读取工作表名称（根据文件格式选择不同的库）
    try:
        is_xls = ext.lower() == '.xls'
        if is_xls:
            # 使用xlrd读取旧版.xls文件
            wb = xlrd.open_workbook(file_path)
            sheet_names = wb.sheet_names()
        else:
            # 使用openpyxl读取.xlsx文件
            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
    except Exception as e:
        os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"无法读取Excel文件: {str(e)}")
    
    # 存储元数据
    asset = {
        'id': file_id,
        'name': saved_name,
        'originalName': file.filename,
        'size': len(content),
        'uploadedAt': datetime.now().isoformat(),
        'sheetNames': sheet_names,
        'path': file_path,
    }
    data_assets[file_id] = asset
    
    return {
        'id': asset['id'],
        'name': asset['name'],
        'originalName': asset['originalName'],
        'size': asset['size'],
        'uploadedAt': asset['uploadedAt'],
        'sheetNames': asset['sheetNames'],
    }


@router.get("")
async def list_assets():
    """获取所有Excel文件资源"""
    return [
        {
            'id': a['id'],
            'name': a['name'],
            'originalName': a['originalName'],
            'size': a['size'],
            'uploadedAt': a['uploadedAt'],
            'sheetNames': a['sheetNames'],
        }
        for a in data_assets.values()
    ]


@router.delete("/{file_id}")
async def delete_asset(file_id: str):
    """删除Excel文件资源"""
    if file_id not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[file_id]
    
    # 删除文件
    if os.path.exists(asset['path']):
        os.remove(asset['path'])
    
    # 删除元数据
    del data_assets[file_id]
    
    return {'message': '删除成功'}


@router.post("/read")
async def read_excel(request: ReadExcelRequest):
    """读取Excel数据"""
    if request.fileId not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[request.fileId]
    file_path = asset['path']
    is_xls = file_path.lower().endswith('.xls')
    
    try:
        if is_xls:
            # 使用xlrd读取.xls文件
            return await _read_excel_xls(file_path, request)
        else:
            # 使用openpyxl读取.xlsx文件
            return await _read_excel_xlsx(file_path, request)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"读取Excel失败: {str(e)}")


async def _read_excel_xlsx(file_path: str, request: ReadExcelRequest):
    """使用openpyxl读取.xlsx文件"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    # 选择工作表
    if request.sheetName:
        if request.sheetName not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"工作表 '{request.sheetName}' 不存在")
        ws = wb[request.sheetName]
    else:
        ws = wb.active
    
    result = None
    result_type = 'unknown'
    
    if request.readMode == 'cell':
        if not request.cellAddress:
            raise HTTPException(status_code=400, detail="单元格模式需要指定cellAddress")
        cell = ws[request.cellAddress]
        result = cell.value
        result_type = 'cell'
    
    elif request.readMode == 'row':
        if request.rowIndex is None:
            raise HTTPException(status_code=400, detail="行模式需要指定rowIndex")
        row_data = []
        for cell in ws[request.rowIndex]:
            row_data.append(cell.value)
        result = row_data
        result_type = 'array'
    
    elif request.readMode == 'column':
        if request.columnIndex is None:
            raise HTTPException(status_code=400, detail="列模式需要指定columnIndex")
        col_data = []
        col_idx = request.columnIndex
        if isinstance(col_idx, str):
            col_idx = openpyxl.utils.column_index_from_string(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            col_data.append(row[0].value)
        result = col_data
        result_type = 'array'
    
    elif request.readMode == 'range':
        if not request.startCell or not request.endCell:
            raise HTTPException(status_code=400, detail="范围模式需要指定startCell和endCell")
        range_data = []
        for row in ws[f"{request.startCell}:{request.endCell}"]:
            row_data = [cell.value for cell in row]
            range_data.append(row_data)
        result = range_data
        result_type = 'matrix'
    
    else:
        raise HTTPException(status_code=400, detail=f"不支持的读取模式: {request.readMode}")
    
    wb.close()
    
    return {'data': result, 'type': result_type}


async def _read_excel_xls(file_path: str, request: ReadExcelRequest):
    """使用xlrd读取.xls文件"""
    wb = xlrd.open_workbook(file_path)
    
    # 选择工作表
    if request.sheetName:
        if request.sheetName not in wb.sheet_names():
            raise HTTPException(status_code=400, detail=f"工作表 '{request.sheetName}' 不存在")
        ws = wb.sheet_by_name(request.sheetName)
    else:
        ws = wb.sheet_by_index(0)
    
    result = None
    result_type = 'unknown'
    
    if request.readMode == 'cell':
        if not request.cellAddress:
            raise HTTPException(status_code=400, detail="单元格模式需要指定cellAddress")
        # 解析单元格地址如 'A1' -> (0, 0)
        col_str, row_str = '', ''
        for c in request.cellAddress:
            if c.isalpha():
                col_str += c
            else:
                row_str += c
        col_idx = _col_letter_to_index(col_str)
        row_idx = int(row_str) - 1
        result = ws.cell_value(row_idx, col_idx)
        result_type = 'cell'
    
    elif request.readMode == 'row':
        if request.rowIndex is None:
            raise HTTPException(status_code=400, detail="行模式需要指定rowIndex")
        row_idx = request.rowIndex - 1  # xlrd从0开始
        row_data = ws.row_values(row_idx)
        result = row_data
        result_type = 'array'
    
    elif request.readMode == 'column':
        if request.columnIndex is None:
            raise HTTPException(status_code=400, detail="列模式需要指定columnIndex")
        col_idx = request.columnIndex
        if isinstance(col_idx, str):
            col_idx = _col_letter_to_index(col_idx)
        else:
            col_idx = col_idx - 1  # xlrd从0开始
        col_data = ws.col_values(col_idx)
        result = col_data
        result_type = 'array'
    
    elif request.readMode == 'range':
        if not request.startCell or not request.endCell:
            raise HTTPException(status_code=400, detail="范围模式需要指定startCell和endCell")
        # 解析范围
        start_col, start_row = _parse_cell_address(request.startCell)
        end_col, end_row = _parse_cell_address(request.endCell)
        range_data = []
        for r in range(start_row, end_row + 1):
            row_data = []
            for c in range(start_col, end_col + 1):
                row_data.append(ws.cell_value(r, c))
            range_data.append(row_data)
        result = range_data
        result_type = 'matrix'
    
    else:
        raise HTTPException(status_code=400, detail=f"不支持的读取模式: {request.readMode}")
    
    return {'data': result, 'type': result_type}


def _col_letter_to_index(col_str: str) -> int:
    """将列字母转换为索引（A=0, B=1, ...）"""
    result = 0
    for c in col_str.upper():
        result = result * 26 + (ord(c) - ord('A') + 1)
    return result - 1


def _parse_cell_address(address: str) -> tuple[int, int]:
    """解析单元格地址，返回(col_idx, row_idx)，从0开始"""
    col_str, row_str = '', ''
    for c in address:
        if c.isalpha():
            col_str += c
        else:
            row_str += c
    return _col_letter_to_index(col_str), int(row_str) - 1


# 提供给执行器使用的函数
def get_asset_path(file_id: str) -> Optional[str]:
    """获取文件路径"""
    if file_id in data_assets:
        return data_assets[file_id]['path']
    return None


def get_asset_by_name(name: str) -> Optional[dict]:
    """通过原始文件名获取资产"""
    for asset in data_assets.values():
        if asset['originalName'] == name:
            return asset
    return None


@router.get("/{file_id}/preview")
async def preview_excel(file_id: str, sheet: Optional[str] = None, max_rows: int = 100, max_cols: int = 50):
    """预览Excel文件数据"""
    if file_id not in data_assets:
        raise HTTPException(status_code=404, detail="文件不存在")
    
    asset = data_assets[file_id]
    file_path = asset['path']
    is_xls = file_path.lower().endswith('.xls')
    
    try:
        if is_xls:
            return _preview_xls(file_path, sheet, max_rows, max_cols)
        else:
            return _preview_xlsx(file_path, sheet, max_rows, max_cols)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"预览失败: {str(e)}")


def _preview_xlsx(file_path: str, sheet_name: Optional[str], max_rows: int, max_cols: int):
    """预览xlsx文件"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise HTTPException(status_code=400, detail=f"工作表 '{sheet_name}' 不存在")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    data = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols), 1):
        row_data = []
        for cell in row:
            val = cell.value
            row_data.append(str(val) if val is not None else '')
        data.append(row_data)
    
    # 获取实际的行列数
    total_rows = ws.max_row or 0
    total_cols = ws.max_column or 0
    
    wb.close()
    
    return {
        'data': data,
        'totalRows': total_rows,
        'totalCols': total_cols,
        'previewRows': len(data),
        'previewCols': max_cols,
    }


def _preview_xls(file_path: str, sheet_name: Optional[str], max_rows: int, max_cols: int):
    """预览xls文件"""
    wb = xlrd.open_workbook(file_path)
    
    if sheet_name:
        if sheet_name not in wb.sheet_names():
            raise HTTPException(status_code=400, detail=f"工作表 '{sheet_name}' 不存在")
        ws = wb.sheet_by_name(sheet_name)
    else:
        ws = wb.sheet_by_index(0)
    
    data = []
    rows_to_read = min(max_rows, ws.nrows)
    cols_to_read = min(max_cols, ws.ncols)
    
    for row_idx in range(rows_to_read):
        row_data = []
        for col_idx in range(cols_to_read):
            val = ws.cell_value(row_idx, col_idx)
            row_data.append(str(val) if val is not None and val != '' else '')
        data.append(row_data)
    
    return {
        'data': data,
        'totalRows': ws.nrows,
        'totalCols': ws.ncols,
        'previewRows': len(data),
        'previewCols': cols_to_read,
    }
